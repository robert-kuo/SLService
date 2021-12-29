import os
import pandas as pd
import datetime as dt
import numpy as np
import sys
import csv, Opt_func

from openpyxl.chart import (LineChart, BarChart, Reference,)
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
# from openpyxl import Workbook

def daterange(date1, date2):
    for n in range((dt.datetime(date2.year, date2.month, date2.day, 0, 0, 0) - dt.datetime(date1.year, date1.month, date1.day, 0, 0, 0)).days + 1):
        yield date1 + dt.timedelta(n)

def abbrev(s):
    abbr = s
    if s == 'Production':
        abbr = 'P'
    elif s == 'Shutdown':
        abbr = 'S'
    elif s == 'No Bootup':
        abbr = 'NB'
    elif s == 'Tunning & Bootup':
        abbr = 'TB'
    elif s == 'Tunning-Production':
        abbr = 'TP'
    return abbr

def Remove_FisrtBlock(lst_blk):
    lst_num = []
    for i in range(lst_blk.shape[0]):
        lst_num.append(i * lst_blk.shape[1])
    nd_tmp = np.reshape(lst_blk, lst_blk.shape[0] * lst_blk.shape[1])
    nd_tmp = np.delete(nd_tmp, lst_num)
    lst_blk = np.reshape(nd_tmp, (lst_blk.shape[0], lst_blk.shape[1] - 1))
    return lst_blk

def Linecount(data_array):
    r_End = data_array.shape[0]
    c_End = data_array.shape[1]

    ret = '當天開機數'
    for i in range(1, c_End):
        if data_array[0, i] != '' and data_array[0, i] != data_array[0, i - 1]:
            n = 0
            for k in range(2, r_End, 3):
                S1 = data_array[0, i]
                m = 0
                v = 0
                while i + m * 4 < c_End and S1 == data_array[0, i + m * 4]:
                    if data_array[k, i + m * 4] == '':
                        sblock = data_array[k + 1, i + m * 4]
                    else:
                        sblock = data_array[k + 2, i + m * 4 + 1]
                    if sblock != '':
                        if str(sblock).find('NB') == -1:
                            v = 1
                        else:
                            if str(sblock).find('NB') >= 0 and str(sblock).find('、') >= 0: v = 1
                    m += 1
                n += v
            ret += ',' + str(n)
            z = 0
            if i + 1 < c_End and data_array[0, i] == data_array[0, i + 1]: z += 3
            if i + 4 < c_End and data_array[0, i] == data_array[0, i + 4]: z += 4
            if i + 8 < c_End and data_array[0, i] == data_array[0, i + 8]: z += 4
            if i + 12 < c_End and data_array[0, i] == data_array[0, i + 12]: z += 4
            i += z
    return ret

# ===== PSR =====
def GetTuneStep(index, block_id, lst_block, df_molds, step_id, lst_bid=None):
    if lst_bid is None: lst_bid = [14, 15, 16, 17, 18, 19, 7, 8, 9, 10, 11, 12, 13]
    n = 0
    zx = 0
    ret = ''
    for z in range(len(lst_bid)):
        if lst_block[index][block_id][lst_bid[z]] != '' and lst_block[index][block_id][lst_bid[z]][:5] != '估計廢料量':
            n += 1
            if n == step_id:
                ret = lst_block[index][block_id][lst_bid[z]]
                zx = z
    p = ret.find('模頭號碼=')
    if p >= 0:
        q = ret.find(',', p + 5)
        if q < 0:
            moldcode = ret[p+5:]
        else:
            moldcode = ret[p+5:q]
        df_tmp = df_molds[df_molds['mold_code'] == moldcode]
        if df_tmp.shape[0] > 0:
            moldno = df_molds.loc[df_tmp.index[0], 'mold_no']
            ret = ret.replace(moldcode, str(moldno) + '(' + moldcode + ')')

    if ret == '': # or ret[:5] == '估計廢料量':
        ret = ',,'
    elif ret[:4] == '調機時間':
        ret = '0.0生產中調機,"' + ret + ',' + lst_block[index][block_id][lst_bid[zx + 1]] + '",'
    else:
        p = ret.find(',')
        if p > 1: ret = ret[:p] + ',"' + ret[p+1:] + '",'
    return ret

def SP_Report(mainpath, taskname, lst_block, df_products, df_orders, df_molds, df_linedata):
    fn = open('sp_report.txt', 'w')
    if os.path.isfile(os.path.join(os.path.join(os.path.join(mainpath, taskname), 'Dataset'), 'lines.csv')):
        df_lines = pd.read_csv(os.path.join(os.path.join(os.path.join(mainpath, taskname), 'Dataset'), 'lines.csv'))
        df_lines['Line Begin'] = df_lines['Line Begin'].astype('datetime64[ns]')
        df_lines.fillna('', inplace=True)
    else:
        df_lines = Opt_func.readlines(df_linedata, df_products)

    lst_lines = Opt_func.sortlist_bynum(df_lines[df_lines['Usable'] == 'YES'].line_name.to_list())  # ['D2', 'Y4', 'Y6', 'Y7', 'Y8', 'Y9', 'Y10']
    lst_SP = [''] * 15
    for i in range(len(lst_lines)):
        ncount = 0
        totwaste = 0
        index = df_lines[df_lines['line_name'] == lst_lines[i]].index[0]
        lst_SP[0] = '生產線:,' + lst_lines[i] + ','
        lst_SP[1] = '排程開始:,' + chr(39) + str(df_lines.loc[index, 'Line Begin'])[5:-3] + ','
        lst_SP[2] = '生產線狀態:,' + df_lines.loc[index, 'Line_Status'] + ','
        lst_SP[3] = '模頭編號:,' + str(df_lines.loc[index, 'mold_no']) + ','
        lst_SP[4] = '設定的製造寬度,' + str(df_lines.loc[index, 'mfg_width']) + ','
        lst_SP[5] = '派工的產品料號,' + df_lines.loc[index, 'part_no'] + ','
        lst_SP[6] = '板材材料,' + df_lines.loc[index, 'material'] + ','
        lst_SP[7] = '生產板材種類,' + df_lines.loc[index, 'type'] + ','
        lst_SP[8] = '結構間距,' + str(df_lines.loc[index, 'lenti_pitch']) + ','
        lst_SP[9] = '結構輪位置,' + df_lines.loc[index, 'roller_position'] + ','
        lst_SP[10] = '設定的產品寬度,' + str(df_lines.loc[index, 'width']) + ','
        lst_SP[11] = '設定的製造厚度,' + str(df_lines.loc[index, 'thickness']) + ','
        lst_SP[12] = '材料的回收料配比,' + df_lines.loc[index, 'composition'] + ','
        lst_SP[13] = '總訂單數,@1,'
        lst_SP[14] = '總廢料量,@2,'
        for v in range(1, Opt_func.Get_blockcount(i, lst_block)):
            block_type = lst_block[i][v][0]
            if lst_block[i][v][5] != '': totwaste += float(lst_block[i][v][5])
            lst_SP[0] += 'Block Type:,' + block_type + ','
            lst_SP[1] += 'Begin:,' + lst_block[i][v][1] + ','
            lst_SP[2] += 'Duration:,' + str(lst_block[i][v][2]) + ' hrs' + ','
            lst_SP[3] += 'Instructions,Parameters,'
            if block_type == 'Tunning & Bootup':
                for k in range(4, 14):
                    lst_SP[k] += GetTuneStep(i, v, lst_block, df_molds, k - 3)
                lst_SP[14] += '估計廢料量,' + '{:3.2f}'.format(int(lst_block[i][v][5] * 100) / 100) + ' kg' + ','
            elif block_type == 'Production':
                ncount += 1
                lst_SP[4] += '訂單編號,' + lst_block[i][v][3] + ','
                lst_SP[5] += '產品編號,' + lst_block[i][v][4] + ','
                # print('xxx', lst_block[i][v][4])
                if lst_block[i][v][4] == '':
                    lst_SP[6] += '產品料號,,'
                else:
                    lst_SP[6] += '產品料號,' + df_products.loc[df_products[df_products['product_code'] == lst_block[i][v][4]].index[0], 'part_no'] + ','
                if lst_block[i][v][3] == '':
                    lst_SP[7] += '需求日期,.'
                else:
                    lst_SP[7] += '需求日期,' + str(df_orders.loc[df_orders[df_orders['order_code'] == lst_block[i][v][3]].index[0], 'not_after'] + dt.timedelta(days=1))[:10] + ','
                for k in range(8, 13):
                    lst_SP[k] += ',,'
                lst_SP[13] += '訂單生產,生產量 = ' + str(lst_block[i][v][6]) + ' (pcs),'
                lst_SP[14] += '估計廢料量,' + '{:3.2f}'.format(int(lst_block[i][v][5] * 100) / 100) + ' kg' + ','
            elif block_type == 'Shutdown':
                lst_SP[4] += '停機(shutdown),' + str(lst_block[i][v][2]) + ' hrs' + ','
                lst_SP[5] += '估計廢料量,' + '{:3.2f}'.format(int(lst_block[i][v][5] * 100) / 100) + ' kg' + ','
                for k in range(6, 15):
                    lst_SP[k] += ',,'
            elif block_type == 'No Bootup':
                lst_SP[4] += 'None,,'
                for k in range(5, 15):
                    lst_SP[k] += ',,'
            elif block_type == 'Tunning-Production':
                ncount += 1
                lst_SP[4] += '訂單編號,' + lst_block[i][v][3] + ','
                lst_SP[5] += '產品編號,' + lst_block[i][v][4] + ','
                lst_SP[6] += '產品料號,' + df_products.loc[df_products[df_products['product_code'] == lst_block[i][v][4]].index[0], 'part_no'] + ','
                lst_SP[7] += '需求日期,' + str(df_orders.loc[df_orders[df_orders['order_code'] == lst_block[i][v][3]].index[0], 'not_after'] + dt.timedelta(days=1))[:10] + ','
                # lst_SP[8] += '生產中調機,' + lst_block[i][v][7] + ','
                # lst_SP[9] += ',' + lst_block[i][v][8] + ','
                for k in range(8, 13):
                    lst_SP[k] += GetTuneStep(i, v, lst_block, df_molds, k - 7, [*range(7, 20)])
                lst_SP[13] += '訂單生產,生產量 = ' + str(lst_block[i][v][6]) + ' (pcs),'
                lst_SP[14] += '估計廢料量,' + '{:3.2f}'.format(int(lst_block[i][v][5] * 100) / 100) + ' kg' + ','

        for k in range(13):
            fn.write(lst_SP[k] + '\n')
        fn.write(lst_SP[13].replace('@1', str(ncount)) + '\n')
        fn.write(lst_SP[14].replace('@2', '{:3.2f}'.format(int(totwaste) / 1000) + '噸') + '\n')
        fn.write('\n')
    fn.close()

def SP_DATA(lst_block, df_lines):
    lst_lines = Opt_func.sortlist_bynum(df_lines[df_lines['Usable'] == 'YES'].line_name.to_list())
    sp_data = np.full((len(lst_lines) + 1, 17), 0, dtype=float)

    for i in range(0, len(lst_lines)):
        m = Opt_func.Get_blockcount(i, lst_block)
        for k in range(1, m):
            if lst_block[i][k][0] == 'Tunning & Bootup':
                sp_data[i, 8] += 1
                sp_data[i, 9] += lst_block[i][k][2]
            elif lst_block[i][k][0] == 'Production':
                sp_data[i, 0] += 1
                sp_data[i, 1] += lst_block[i][k][2]
                sp_data[i, 2] += lst_block[i][k][6]
                sp_data[i, 3] += lst_block[i][k][5]
            elif lst_block[i][k][0] == 'Shutdown':
                sp_data[i, 10] += 1
                sp_data[i, 11] += lst_block[i][k][2]
                sp_data[i, 12] += lst_block[i][k][5]
            elif lst_block[i][k][0] == 'No Bootup':
                sp_data[i, 13] += 1
                sp_data[i, 14] += lst_block[i][k][2]
            else:
                sp_data[i, 4] += 1
                sp_data[i, 5] += lst_block[i][k][2]
                sp_data[i, 6] += lst_block[i][k][6]
                sp_data[i, 7] += lst_block[i][k][5]
    for i in range(0, len(lst_lines)):
        sp_data[i, 15] = sp_data[i, 0] + sp_data[i, 4] + sp_data[i, 8] + sp_data[i, 10] + sp_data[i, 13]
        sp_data[i, 16] = sp_data[i, 1] + sp_data[i, 5] + sp_data[i, 9] + sp_data[i, 11] + sp_data[i, 14]

    for k in range(sp_data.shape[1]):
        sum = 0
        for i in range(sp_data.shape[0] - 1):
            sum += sp_data[i, k]
            sp_data[sp_data.shape[0] - 1, k] = sum
    return sp_data

def DemandBalanceSheet(demand_start, demand_end, lst_weekday, lst_block, df_orderdata, df_products, df_lines):
    columns = ['product_code', 'part_no', 'material', 'type', 'width', 'length', 'height', 'density', '數量項目']
    lcount = len(columns)
    rcount = (demand_end - demand_start).days + 1
    lst_week = [''] * lcount
    lst_one = []

    for i in range(rcount):
        t = demand_start + dt.timedelta(days=i)
        if t.day == 1: lst_one.append(lcount + i + len(lst_one) * 2)
        columns.append(str(t)[:10])
        lst_week.append(lst_weekday[t.weekday()])

    df_balancedata = pd.DataFrame(columns=columns)
    df_balancedata.loc[0] = lst_week

    for i in range(df_products.shape[0]):
        lst_data = [''] * (lcount - 1) + ['Demand'] + [''] * rcount
        df_balancedata.loc[df_balancedata.shape[0]] = lst_data
        lst_data = [df_products.loc[i, 'product_code'], df_products.loc[i, 'part_no'], df_products.loc[i, 'material'],
                    df_products.loc[i, 'type'], df_products.loc[i, 'width'], df_products.loc[i, 'length'],
                    df_products.loc[i, 'height'], df_products.loc[i, 'density']] + ['On hand Stock'] + [''] * rcount
        df_balancedata.loc[df_balancedata.shape[0]] = lst_data
        lst_data = [''] * (lcount - 1) + ['MFG output'] + [''] * rcount
        df_balancedata.loc[df_balancedata.shape[0]] = lst_data

    # Demand
    lst_stock_code = []
    lst_stock_qty = []
    for i in range(df_orderdata.shape[0]):
        index = df_balancedata[df_balancedata['product_code'] == df_orderdata.loc[i, 'product_code']].index[0]
        cid = df_orderdata.columns.get_loc('on_hand_stock')
        lst_stock_code.append(df_orderdata.loc[i, 'product_code'])
        lst_stock_qty.append(df_orderdata.loc[i, 'on_hand_stock'])
        for k in range(cid + 1, len(df_orderdata.columns)):
            df_balancedata.iloc[index - 1, lcount + k - cid - 1] = df_orderdata.iloc[i, k]

    # MFG output
    lst_lines = Opt_func.sortlist_bynum(df_lines[df_lines['Usable'] == 'YES'].line_name.to_list())
    for i in range(len(lst_lines)):
        for v in range(1, Opt_func.Get_blockcount(i, lst_block)):
            block_type = lst_block[i][v][0]
            if block_type == 'Production' or block_type == 'Tunning-Production':
                duration = lst_block[i][v][2]
                tune_dura = 0
                if lst_block[i][v][7] == '調機時間=1.0小時': tune_dura = 1
                if lst_block[i][v][7] == '調機時間=0.5小時': tune_dura = 0.5
                duration -= tune_dura

                qty = lst_block[i][v][6]
                remain = qty
                t1 = dt.datetime.strptime(lst_block[i][v][1], '%Y-%m-%d %H:%M:%S') + dt.timedelta(hours=tune_dura)
                t2 = t1 + dt.timedelta(hours=duration)
                for xt in daterange(t1, t2):
                    if xt == t1:
                        t3 = t1
                    else:
                        t3 = dt.datetime(xt.year, xt.month, xt.day, 0, 0, 0)
                    t4 = t3 + dt.timedelta(days=1)
                    t4 = dt.datetime(t4.year, t4.month, t4.day, 0, 0, 0)
                    if t4 > t2: t4 = t2
                    if t3 != t4:
                        qty_day = df_products.loc[
                            df_products[df_products['product_code'] == lst_block[i][v][4]].index[0], 'throughput']
                        if (t4 - t3).days * 86400 + (t4 - t3).seconds != 86400:
                            if t4.hour == 0 and t4.minute == 0 and t4 != t2:
                                qty_day = round((t4 - t3).days * 86400 + (t4 - t3).seconds * qty_day / 86400)
                            else:
                                qty_day = remain
                        if remain < qty_day: qty_day = remain

                        # qty_day = round(((t4 - t3).days * 86400 + (t4 - t3).seconds) * qty / (3600 * duration))
                        # if abs(remain - qty_day) < 10: qty_day = remain

                        remain -= qty_day
                        # print(lst_block[i][v][4], t3, t4, duration, qty_day, remain, lst_block[i][v][7])
                        if t3 < demand_start: t3 = demand_start
                        index = df_balancedata[df_balancedata['product_code'] == lst_block[i][v][4]].index[0]
                        # print(qty_day)
                        if df_balancedata.loc[index + 1, str(t3)[:10]] == '':
                            df_balancedata.loc[index + 1, str(t3)[:10]] = qty_day
                        else:
                            df_balancedata.loc[index + 1, str(t3)[:10]] += qty_day

    # On hand Stock
    lst_Product_negStock = []
    for i in range(2, df_balancedata.shape[0], 3):
        index = Opt_func.list_index(lst_stock_code, df_balancedata.loc[i, 'product_code'])
        if index >= 0:
            stock_qty = lst_stock_qty[index]
            mfg_qty = 0
            for k in range(lcount, lcount + rcount):
                demnad_qty = 0 if df_balancedata.iloc[i - 1, k] == '' else df_balancedata.iloc[i - 1, k]
                stock_qty += mfg_qty - demnad_qty
                df_balancedata.iloc[i, k] = stock_qty
                if stock_qty < 0 and Opt_func.list_index(lst_Product_negStock,
                                                df_balancedata.loc[i, 'product_code']) < 0: lst_Product_negStock.append(
                    df_balancedata.loc[i, 'product_code'])
                mfg_qty = 0 if df_balancedata.iloc[i + 1, k] == '' else df_balancedata.iloc[i + 1, k]
    df_negStock = df_balancedata[df_balancedata['product_code'].isin(lst_Product_negStock)]

    for i in range(len(lst_one)):
        sTotal = 'Total-' + '{0:02d}'.format((demand_start.month + i - 1) % 12 + 1)
        sTTL = 'TTL-' + '{0:02d}'.format((demand_start.month + i - 1) % 12 + 1)
        df_balancedata.insert(lst_one[i], sTotal, '')
        df_balancedata.insert(lst_one[i] + 1, sTTL, '')
        df_balancedata.loc[0, sTotal] = '重量(Ton)'
        df_balancedata.loc[0, sTTL] = 'pcs'

        i1 = lst_one[i] + 2
        i2 = len(df_balancedata.columns) if i == len(lst_one) - 1 else lst_one[i + 1]

        for v in range(2, df_balancedata.shape[0], 3):
            sum1 = 0
            sum2 = 0
            for k in range(i1, i2):
                sum1 += 0 if df_balancedata.iloc[v - 1, k] == '' else df_balancedata.iloc[v - 1, k]
                sum2 += 0 if df_balancedata.iloc[v + 1, k] == '' else df_balancedata.iloc[v + 1, k]
            df_balancedata.iloc[v - 1, i1 - 1] = sum1
            df_balancedata.iloc[v + 1, i1 - 1] = sum2
            index = Opt_func.list_index(lst_stock_code, df_balancedata.loc[v, 'product_code'])
            df_balancedata.iloc[v, i1 - 1] = lst_stock_qty[index]

            w = df_balancedata.loc[v, 'width'] * df_balancedata.loc[v, 'length'] * df_balancedata.loc[v, 'height'] * \
                df_balancedata.loc[v, 'density'] / (1000000 * 1000)
            for k in range(-1, 2):
                df_balancedata.iloc[v + k, i1 - 2] = '{:0.1f}'.format(w * df_balancedata.iloc[v + k, i1 - 1])

    return df_balancedata, df_negStock

def DrawOnStock(df_negStock, writer, thin):
    lst_delete = []
    for i in range(df_negStock.shape[0]):
        bDelete = True
        for k in range(9, df_negStock.shape[1]):
            if df_negStock.iloc[i, k] < 0:
                bDelete = False
                break
        if bDelete: lst_delete.append(df_negStock.index[i])
    if len(lst_delete) > 0: df_negStock.drop(lst_delete, inplace=True)

    workbook = writer.book
    worksheet = workbook.create_sheet('逾期生產的產品')

    worksheet.cell(row=1, column=1).value = 'product_code'
    worksheet.cell(row=2, column=1).value = 'part_no'
    worksheet.cell(row=3, column=1).value = '最終不足的生產量'
    worksheet.cell(row=4, column=1).value = '庫存不足的天數'
    worksheet.cell(row=5, column=1).value = '最大的不足量'
    worksheet.cell(row=6, column=2).value = 'On hand Stock'
    worksheet.cell(row=7, column=1).value = '日期'
    worksheet.cell(row=6, column=2).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.merge_cells(start_row=6, start_column=2, end_row=6, end_column=1 + df_negStock.shape[0])
    for i in range(df_negStock.shape[0]):
        worksheet.cell(row=1, column=i + 2).value = df_negStock.iloc[i, 0]
        worksheet.cell(row=7, column=i + 2).value = df_negStock.iloc[i, 0]
        worksheet.cell(row=2, column=i + 2).value = df_negStock.iloc[i, 1]
        worksheet.cell(row=3, column=i + 2).value = df_negStock.iloc[i, df_negStock.shape[
            1] - 1]  # df_order_lost[df_order_lost['product_code'] == df_negStock.iloc[i, 0]]['quantity'].sum()
        n = 0
        min_stock = 10000
        for k in range(9, df_negStock.shape[1]):
            if df_negStock.iloc[i, k] < 0:
                n += 1
                if min_stock > df_negStock.iloc[i, k]: min_stock = df_negStock.iloc[i, k]
        worksheet.cell(row=4, column=i + 2).value = n
        worksheet.cell(row=5, column=i + 2).value = min_stock

    for k in range(9, df_negStock.shape[1]):
        worksheet.cell(row=k - 1, column=1).value = df_negStock.columns[k]
        for i in range(df_negStock.shape[0]):
            worksheet.cell(row=k - 1, column=i + 2).value = df_negStock.iloc[i, k]
    for i in range(df_negStock.shape[1] - 2):
        for k in range(df_negStock.shape[0] + 1):
            thecell = worksheet[chr(65 + k) + str(i + 1)]
            thecell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    values = Reference(worksheet, min_col=2, min_row=7, max_col=df_negStock.shape[0] + 1,
                       max_row=df_negStock.shape[1] - 3)
    chart = LineChart()
    chart.add_data(values, titles_from_data=True)
    dates_title = Reference(worksheet, min_col=1, min_row=8, max_row=df_negStock.shape[1] - 2)
    chart.set_categories(dates_title)

    chart.title = ' 逾期生產的庫存趨勢圖 '
    chart.x_axis.title = ' Date '
    chart.y_axis.title = ' Qutatity '
    chart.width = 21
    chart.height = 14
    worksheet.add_chart(chart, chr(67 + df_negStock.shape[0]) + '1')

def ImpSector_Sheet(lst_blk, lst_weekday, max_linearray, begin_day, end_day, maxline_count, df_lines):
    lst_blk = Remove_FisrtBlock(lst_blk)
    lst_type = ['NB Sector', 'Instatnt Sector', 'NewP Sector']
    lst_lines = Opt_func.sortlist_bynum(df_lines[df_lines['Usable'] == 'YES'].line_name.to_list())  # ['D2', 'Y4', 'Y6', 'Y7', 'Y8', 'Y9', 'Y10']

    min_day = lst_blk[0][0][1][:10]
    for i in range(1, len(lst_blk)):
        t = lst_blk[i][0][1][:10]
        if t < min_day: min_day = t

    tx = dt.datetime.strptime(min_day, '%Y-%m-%d')
    x = Opt_func.getdays(min_day, end_day)
    lst_column = [1] * (x + 1)  # column count for all lines
    lst_date = []
    for i in range(len(lst_column)):
        lst_date.append(tx + dt.timedelta(days=i))

    imp_sector = np.zeros([len(lst_lines), 3, 4], dtype=float, order='C')
    lst_imp = []  # imp sector time for each lines
    lst_ibk = []  # imp block id for each lines
    lst_col = []  # column count for each lines

    # find imp sector
    for i in range(len(lst_blk)):
        lst_data = [dt.datetime.strptime(lst_blk[i][0][1], '%Y-%m-%d %H:%M:%S')]  # [dt.datetime(tx.year, tx.month, tx.day, tx.hour, tx.minute)]
        lst_blockid = [0]
        lst_coldata = [1] * (x + 1)
        lst_coldata[Opt_func.getdays(min_day, lst_blk[i][0][1][:10])] = 4

        pid = -1
        for v in range(Opt_func.Get_blockcount(i, lst_blk)):
            if len(lst_blk[i][v]) > 0:
                block_name = lst_blk[i][v][0]
                if Opt_func.list_index(['No Bootup', 'Production', 'Tunning-Production'], block_name) >= 0:
                    block_time = dt.datetime.strptime(lst_blk[i][v][1], '%Y-%m-%d %H:%M:%S') + dt.timedelta(hours=lst_blk[i][v][2])
                    lst_data.append(block_time)
                    lst_blockid.append(v + 1)
                    nid = Opt_func.list_index(lst_date, dt.datetime(block_time.year, block_time.month, block_time.day))
                    lst_coldata[nid] = 4 if nid != pid else (lst_coldata[nid] + 4)
                    pid = nid

        lst_col.append(lst_coldata)
        lst_imp.append(lst_data)
        lst_ibk.append(lst_blockid)

    c_sum = 0
    for v in range(len(lst_column)):
        c_max = 0
        for i in range(len(lst_col)):
            if c_max < lst_col[i][v]: c_max = lst_col[i][v]
        lst_column[v] = c_max
        c_sum += c_max

    np.set_printoptions(threshold=sys.maxsize)
    data_array = np.full((len(lst_blk) * 3 + 2, c_sum + 1), '', dtype=object)

    # print(data_array.shape)
    # output
    # title: date and weekday
    lst_tmp = []
    p = 1
    for i in range(len(lst_column)):
        for v in range(lst_column[i]):
            data_array[0, p] = str(lst_date[i].month) + '/' + str(lst_date[i].day)
            data_array[1, p] = lst_weekday[lst_date[i].weekday()]
            lst_tmp.append(data_array[0, p])
            p += 1

    # print('lst_imp', len(lst_imp))
    for i in range(len(lst_imp)):
        # line name
        data_array[i * 3 + 2, 0] = lst_lines[i]
        data_array[i * 3 + 3, 0] = lst_lines[i]
        data_array[i * 3 + 4, 0] = lst_lines[i]

        n = Opt_func.Get_blockcount(i, lst_blk)
        for v in range(len(lst_imp[i])):
            t1 = lst_imp[i][v]
            id1 = Opt_func.list_index(lst_tmp, str(t1.month) + '/' + str(t1.day))
            if id1 >= 0:
                start_hour = str(t1.hour) + ':' + str(t1.minute)

                # summary imp sector information
                sub_waste = 0
                sub_hours = 0
                sub_block = ''
                sub_sector = ''
                sub_order = ''
                sub_qty = 0
                if v < len(lst_imp[i]) - 1:
                    end_id = lst_ibk[i][v + 1]
                else:
                    end_id = n
                if lst_ibk[i][v] == end_id:
                    t1 = dt.datetime.strptime(end_day, '%Y-%m-%d')
                else:
                    t1 = dt.datetime.strptime(lst_blk[i][lst_ibk[i][v]][1], '%Y-%m-%d %H:%M:%S')
                t1 = dt.datetime(t1.year, t1.month, t1.day)
                t2 = dt.datetime.strptime(lst_blk[i][end_id - 1][1], '%Y-%m-%d %H:%M:%S') + dt.timedelta(hours=lst_blk[i][end_id - 1][2])
                bk_all_days = (t2 - t1).days

                for w in range(lst_ibk[i][v], end_id):
                    block_name = abbrev(lst_blk[i][w][0])
                    if sub_order == '': sub_order = lst_blk[i][w][3]
                    if lst_blk[i][w][6] != '': sub_qty += lst_blk[i][w][6]
                    if lst_blk[i][w][5] != '': sub_waste += lst_blk[i][w][5]
                    if lst_blk[i][w][2] != '': sub_hours += lst_blk[i][w][2]
                    sub_block += block_name + '(' + str(lst_blk[i][w][2]) + 'h)、'
                    sub_sector += block_name + ','

                # fill data into imp sector
                w = 1
                while w > 0:
                    if data_array[i * 3 + 2, id1 + w] == '':
                        data_array[i * 3 + 2, id1 + w] = '起始時間'
                        data_array[i * 3 + 2, id1 + w + 1] = start_hour
                        data_array[i * 3 + 2, id1 + w + 2] = '訂單編號'
                        data_array[i * 3 + 2, id1 + w + 3] = sub_order
                        data_array[i * 3 + 3, id1 + w] = '生產量'
                        data_array[i * 3 + 3, id1 + w + 1] = sub_qty
                        data_array[i * 3 + 3, id1 + w + 2] = '估計廢料量'
                        data_array[i * 3 + 3, id1 + w + 3] = round(sub_waste, 2)
                        data_array[i * 3 + 4, id1 + w] = '連續區塊類別'
                        data_array[i * 3 + 4, id1 + w + 1] = sub_block[:-1]
                        # print('aaa', i, id1, w, str(t1.month) + '/' + str(t1.day))
                        # print(data_array[i * 3 + 2, id1 + w], data_array[i * 3 + 2, id1 + w + 1],data_array[i * 3 + 2, id1 + w + 2], data_array[i * 3 + 2, id1 + w + 3])
                        # print(data_array[i * 3 + 3, id1 + w], data_array[i * 3 + 3, id1 + w + 1],data_array[i * 3 + 3, id1 + w + 2], data_array[i * 3 + 3, id1 + w + 3])
                        # print(data_array[i * 3 + 4, id1 + w], data_array[i * 3 + 4, id1 + w + 1])
                        w = 0
                    else:
                        w += 4

                # fill block name not in imp_sector
                for w in range(lst_ibk[i][v], end_id):
                    t1 = dt.datetime.strptime(lst_blk[i][w][1], '%Y-%m-%d %H:%M:%S')
                    t1 = dt.datetime(t1.year, t1.month, t1.day)
                    t2 = dt.datetime.strptime(lst_blk[i][w][1], '%Y-%m-%d %H:%M:%S') + dt.timedelta(hours=lst_blk[i][w][2])
                    for z in range(1 if w == lst_ibk[i][v] and bk_all_days == 0 else (t2 - t1).days):
                        t3 = t1 + dt.timedelta(days=z)
                        block_name = abbrev(lst_blk[i][w][0])
                        id2 = Opt_func.list_index(lst_tmp, str(t3.month) + '/' + str(t3.day))
                        id3 = Opt_func.list_index(lst_date, dt.datetime(t3.year, t3.month, t3.day))
                        for x in range(lst_column[id3]):
                            if data_array[i * 3 + 3, id2 + 1 + x] == '':
                                data_array[i * 3 + 3, id2 + 1 + x] = block_name
                                # print('bbb', i, id2, x, data_array[i * 3 + 3, id2 + 1 + x])

                if sub_sector == 'S,NB,' or sub_sector == 'NB,':
                    sec_id = 0
                elif sub_sector == 'TP,' or sub_sector == 'P,':
                    sec_id = 1
                elif sub_sector == 'TB,P,' or sub_sector == 'S,TB,P,':
                    sec_id = 2
                else:
                    sec_id = -1
                if sec_id != -1:
                    imp_sector[i, sec_id, 0] += 1
                    imp_sector[i, sec_id, 1] += sub_qty
                    imp_sector[i, sec_id, 2] += sub_waste
                    imp_sector[i, sec_id, 3] += sub_hours

    # for i in range(2, data_array.shape[0]):
    #     for v in range(0, 100):
    #         print('ccc', i, v, data_array[i, v])

    # output ndarray to text
    fn = open('imp_sector.txt', 'w')
    lst_PHData = []
    for i in range(2):
        s = data_array[i, 0]
        for v in range(1, data_array.shape[1]):
            if data_array[i, v] != data_array[i, v - 1]: s += ',' + str(data_array[i, v])
        if i == 0: lst_PHData.append('日期' + s)
        fn.write(s + '\n')

    lst_ordercount = []
    for i in range(2, data_array.shape[0], 3):
        s = data_array[i, 0]
        v = 1

        n_loop = 0
        while v < data_array.shape[1]:
            n_Order = 0
            if v == data_array.shape[1] - 1 and data_array[0, v] != data_array[0, v - 1] or v < data_array.shape[1] - 1 and data_array[0, v] != data_array[0, v - 1] and data_array[0, v] != data_array[0, v + 1]:
                s += ',' + str(data_array[i + 1, v])
            else:
                x = -1
                if v + 3 < data_array.shape[1] and data_array[0, v] == data_array[0, v + 3]: x = 0
                if v + 7 < data_array.shape[1] and data_array[0, v] == data_array[0, v + 7]: x = 1
                if v + 11 < data_array.shape[1] and data_array[0, v] == data_array[0, v + 11]: x = 2
                if v + 15 < data_array.shape[1] and data_array[0, v] == data_array[0, v + 15]: x = 3
                if v + 19 < data_array.shape[1] and data_array[0, v] == data_array[0, v + 19]: x = 4
                if x >= 0:
                    sx = ''
                    for z in range(x + 1):
                        sz = str(data_array[i, v + z * 4])
                        # print(data_array[i, v + z * 4], data_array[i, v + z * 4 + 1], data_array[i, v + z * 4 + 2], data_array[i, v + z * 4 + 3], 'zzz')
                        # print(data_array[i+1, v + z * 4], data_array[i+1, v + z * 4 + 1], data_array[i+1, v + z * 4 + 2], data_array[i+1, v + z * 4 + 3], 'zzz')
                        # print(data_array[i+2, v + z * 4], data_array[i+2, v + z * 4 + 1], data_array[i+2, v + z * 4 + 2], data_array[i+2, v + z * 4 + 3], 'zzz')
                        if sz == '' or sz == '0:0':
                            if z == 0: sx = str(data_array[i + 1, v + z * 4])
                        else:
                            if sx != '': sx += '%%'
                            sx += '1. ' + sz + ':' + str(data_array[i, v + z * 4 + 1]) + '%%'
                            sx += '2. ' + str(data_array[i, v + z * 4 + 2]) + ':' + ('' if str(data_array[i, v + z * 4 + 3]) == '起始時間' else str(data_array[i, v + z * 4 + 3])) + '%%'
                            sx += '3. ' + str(data_array[i + 1, v + z * 4]) + ':' + str(data_array[i + 1, v + z * 4 + 1]) + '%%'
                            sx += '4. ' + str(data_array[i + 1, v + z * 4 + 2]) + ':' + ('' if str(data_array[i + 1, v + z * 4 + 3]) == '生產量' else str(data_array[i + 1, v + z * 4 + 3])) + '%%'
                            sx += '5. ' + str(data_array[i + 2, v + z * 4]) + ':' + str(data_array[i + 2, v + z * 4 + 1])
                            if str(data_array[i + 2, v + z * 4 + 1]).find('P') >= 0: n_Order += 1
                        if sx == '0': sx = str(data_array[i + 1, v + z * 4 + 3])
                    # print('ddd', i, v, x, sx)
                    s += ',' + sx
                    v += 3 + 4 * x
            if len(lst_ordercount) <= n_loop:
                lst_ordercount.append(n_Order)
            else:
                lst_ordercount[n_loop] += n_Order
            v += 1
            n_loop += 1
        fn.write(s + '\n')

    lst_PHs =  Opt_func.GetEvery_PHs(lst_blk, len(lst_ordercount), df_lines)
    fn.write(Linecount(data_array) + '\n')
    s = '啟動的訂單數'
    for i in range(len(lst_ordercount)):
        s += ',' + str(lst_ordercount[i])
    fn.write(s + '\n')
    s = '實際生產時數'
    for i in range(len(lst_PHs)):
        s += ',' + str(lst_PHs[i])
    lst_PHData.append(s)
    fn.write(s + '\n')

    lst_linearray = [0] * 214
    max_dataarray = [sum(i) for i in zip(max_linearray, lst_linearray)]

    v = Opt_func.getdays(begin_day, min_day)
    s = '預定的彈性生產日'
    s1 = '彈性增加的工時'
    for i in range(v, len(max_dataarray)):
        s += ',@' if max_dataarray[i] > maxline_count else ','
        if lst_PHs[i - v] > maxline_count * 24:
            s1 += ',' + str(lst_PHs[i - v] - maxline_count * 24)
        else:
            s1 += ','
    fn.write(s + '\n')
    fn.write(s1 + '\n')
    lst_PHData.append(s1)
    fn.write('\n')
    fn.close()

    # Sector summary
    columns = ['Line', 'Sector type', 'Sector數量', '總生產量', '總廢料量', '總時數', ]
    df_sector = pd.DataFrame(columns=columns)
    n = 0
    imp_totlines = np.zeros([len(lst_type) + 1, 4], dtype=float, order='C')
    for i in range(len(lst_lines)):
        imp_total = np.zeros([4], dtype=float, order='C')
        for v in range(len(lst_type)):
            df_sector.loc[n] = [lst_lines[i], lst_type[v], int(imp_sector[i, v, 0]), f'{int(imp_sector[i, v, 1]):,}', '{:3.3f}'.format(int(imp_sector[i, v, 2]) / 1000), imp_sector[i, v, 3]]
            for k in range(4):
                imp_total[k] += imp_sector[i, v, k]
                imp_totlines[v, k] += imp_sector[i, v, k]
            n += 1
        df_sector.loc[n] = [lst_lines[i], 'All Types', int(imp_total[0]), f'{int(imp_total[1]):,}', '{:3.3f}'.format(int(imp_total[2]) / 1000), imp_total[3]]
        n += 1
    for v in range(len(lst_type)):
        df_sector.loc[n + v] = ['All Lines', lst_type[v], int(imp_totlines[v, 0]), f'{int(imp_totlines[v, 1]):,}', '{:3.3f}'.format(int(imp_totlines[v, 2]) / 1000), imp_totlines[v, 3]]
        for k in range(4):
            imp_totlines[len(lst_type), k] += imp_totlines[v, k]
    df_sector.loc[n + v + 1] = ['All Lines', 'All Types', int(imp_totlines[len(lst_type), 0]), f'{int(imp_totlines[len(lst_type), 1]):,}', '{:3.3f}'.format(int(imp_totlines[len(lst_type), 2]) / 1000), imp_totlines[len(lst_type), 3]]
    # pretty_print(df_sector)

    # Line Ratio
    columns = ['產線別', '派工數(訂單數)', '生產量(片數)', '總廢料量(噸)', '廢料產生率(kg/pcs)', '生產時數', '總排程時數', '產能利用率']
    df_LRation = pd.DataFrame(columns=columns)
    for i in range(len(lst_blk) + 1):
        w_hour = df_sector.loc[i * 4 + 3, '總時數']
        p_hour = df_sector.loc[i * 4 + 1, '總時數'] + df_sector.loc[i * 4 + 2, '總時數']
        df_LRation.loc[i] = [df_sector.loc[i * 4, 'Line'], df_sector.loc[i * 4 + 1, 'Sector數量'] + df_sector.loc[i * 4 + 2, 'Sector數量'],
                             df_sector.loc[i * 4 + 3, '總生產量'], df_sector.loc[i * 4 + 3, '總廢料量'], '{:3.3f}'.format(0 if float(df_sector.loc[i * 4 + 3, '總生產量'].replace(',', '')) == 0 else float(
                             df_sector.loc[i * 4 + 3, '總廢料量']) * 1000 / float(df_sector.loc[i * 4 + 3, '總生產量'].replace(',', ''))), p_hour, w_hour, '{:0.2%}'.format(p_hour / w_hour)]

    # pretty_print(df_LRation)
    return df_sector, df_LRation, lst_PHData

def Output_Report(mainpath, taskname, output_file, ordercsv_file, data_file, df_products, df_orders, df_lines, df_molds, df_orderdata, df_linedata, ths_index, Stage_name, maxline_count, Learning_No, Learning_max, episode, epi_max, demand_start, demand_end, begin_day, end_day, extra_PHs, tot_waste, lst_weekday, lst_lossorders, max_linearray):
    writer = pd.ExcelWriter(output_file, engine='openpyxl')

    lst_lines = Opt_func.sortlist_bynum(df_lines[df_lines['Usable'] == 'YES'].line_name.to_list())
    workbook = writer.book
    worksheet = workbook.create_sheet('排程演練紀錄')

    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 20
    df_order_result = pd.read_csv(ordercsv_file)
    df_order_done = df_order_result[df_order_result['O_Status'] == 'Done']
    df_order_lost = df_order_result[df_order_result['O_Status'] == 'LOST']
    df_order_lost.reset_index(inplace=True)
    df_order_lost.drop(columns=[df_order_lost.columns[0], df_order_lost.columns[1]], axis=1, inplace=True)

    # worksheet = writer.sheets['排程演練紀錄']
    worksheet.cell(row=1, column=1).value = '排程任務名稱'
    worksheet.cell(row=1, column=2).value = ':'
    worksheet.cell(row=1, column=3).value = 'Darwin7-12月排程'
    worksheet.cell(row=2, column=1).value = '學習演練的階段'
    worksheet.cell(row=2, column=2).value = ':'
    worksheet.cell(row=2, column=3).value = (('Trial' + str(ths_index + 1)) if Stage_name == 'Trial Stage' else 'Target') + ' Practice'

    worksheet.cell(row=3, column=1).value = 'Learning No.'
    worksheet.cell(row=3, column=2).value = Learning_No - 1  # int(learndata[6])
    worksheet.cell(row=3, column=3).value = 'PLT=' + str(Learning_max)
    worksheet.cell(row=4, column=1).value = 'Episode'
    worksheet.cell(row=4, column=2).value = episode - 1  # df_learn.loc[int(learndata[6]) - 1, 'episode count']
    worksheet.cell(row=4, column=3).value = 'Maximum Episode=' + str(epi_max)
    worksheet.cell(row=5, column=1).value = 'Done Orders'
    worksheet.cell(row=5, column=2).value = df_order_done.shape[0]
    worksheet.cell(row=5, column=3).value = 'Waiting Orders = ' + str(df_order_result.shape[0])
    worksheet.cell(row=6, column=1).value = 'Production'
    worksheet.cell(row=6, column=2).value = df_order_done['quantity'].sum()
    worksheet.cell(row=6, column=3).value = '(pcs)'
    worksheet.cell(row=7, column=1).value = 'Production Hours'
    worksheet.cell(row=7, column=2).value = df_order_done['Production_Hours'].sum()
    worksheet.cell(row=7, column=3).value = '(hrs)'
    worksheet.cell(row=8, column=1).value = 'Extra Production Hours'
    worksheet.cell(row=8, column=2).value = extra_PHs
    worksheet.cell(row=8, column=3).value = '(hrs)'
    worksheet.cell(row=9, column=1).value = 'Waste'
    worksheet.cell(row=9, column=2).value = tot_waste  # float(learndata[2])
    worksheet.cell(row=9, column=3).value = '(tons)'
    worksheet.cell(row=10, column=1).value = 'Loss Orders'
    worksheet.cell(row=10, column=2).value = len(lst_lossorders)  # len(lst_lossdata)
    worksheet.cell(row=10, column=3).value = '(ea)'
    worksheet.cell(row=11, column=1).value = 'Pending Production'
    worksheet.cell(row=11, column=2).value = df_order_lost['quantity'].sum()
    worksheet.cell(row=11, column=3).value = '(pcs)'
    worksheet.cell(row=12, column=1).value = 'Unsatisfied PHs'
    worksheet.cell(row=12, column=2).value = df_order_lost['Production_Hours'].sum()
    worksheet.cell(row=12, column=3).value = '(hrs)'

    # draw table border
    thin = Side(border_style='thin', color='000000')
    for i in range(3):
        for k in range(12):
            thecell = worksheet[chr(65 + i) + str(k + 1)]
            thecell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    if df_order_lost.shape[0] > 0:
        df_order_lost.to_excel(writer, sheet_name='掉失的訂單組', startrow=0, startcol=0, index=False)
        df_trendlost = Opt_func.OrderLost_Trendchart(df_order_lost, df_lines, end_day)
        df_trendlost.to_excel(writer, sheet_name='掉失的訂單組', startrow=df_order_lost.shape[0] + 2, startcol=0, index=False)

        worksheet = writer.sheets['掉失的訂單組']
        values = Reference(worksheet, min_col=2, min_row=df_order_lost.shape[0] + 3, max_col=df_trendlost.shape[1], max_row=df_order_lost.shape[0] + df_trendlost.shape[0] + 3)
        chart = LineChart()
        chart.add_data(values, titles_from_data=True)
        dates_title = Reference(worksheet, min_col=1, min_row=df_order_lost.shape[0] + 4, max_row=df_order_lost.shape[0] + df_trendlost.shape[0] + 3)
        chart.set_categories(dates_title)

        chart.title = ' 掉失的訂單的生產線時數需求趨勢圖 '
        chart.x_axis.title = ' Date '
        chart.y_axis.title = ' Request Hours '
        chart.width = 17
        chart.height = 12
        worksheet.add_chart(chart, chr(66 + df_trendlost.shape[1]) + str(df_order_lost.shape[0] + 3))

    # arr = np.array(lst_bestresult)
    # np.save('data_array_0913_01.npy', arr)
    lst_block = np.load(data_file + '.npy', allow_pickle=True)

    # 派工指示區塊表
    SP_Report(mainpath, taskname, lst_block, df_products, df_orders, df_molds, df_linedata)
    worksheet = workbook.create_sheet('派工指示區塊表')
    line_count = 0
    with open('sp_report.txt') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            line_count += 1
            for i in range(len(row)):
                worksheet.cell(row=line_count, column=i + 1).value = row[i]
                if i > 1:
                    worksheet.cell(row=line_count, column=i + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                else:
                    worksheet.cell(row=line_count, column=i + 1).font = Font(color='0000FF')
                if row[i] == 'Block Type:' or row[i] == 'Begin:' or row[i] == 'Duration:' or row[i] == 'Instructions' or row[i] == 'Parameters':
                    worksheet.cell(row=line_count, column=i + 1).fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD',                                                                                    fill_type='solid')

    sp_data = SP_DATA(lst_block, df_lines)
    lst_title = ['區塊別', 'Production', '', '', '', 'Tunning-Production', '', '', '', 'Tunning & Bootup', '', 'Shutdown', '', '', 'No Bootup', '', 'All Types', '']
    lst_merge = [0, 3, 0, 0, 0, 3, 0, 0, 0, 1, 0, 2, 0, 0, 1, 0, 1, 0]
    for i in range(len(lst_title)):
        worksheet.cell(row=line_count + 2, column=i + 1).value = lst_title[i]
        worksheet.cell(row=line_count + 2, column=i + 1).alignment = Alignment(horizontal='center', vertical='center')
        if lst_merge[i] > 0: worksheet.merge_cells(start_row=line_count + 2, start_column=i + 1, end_row=line_count + 2,
                                                   end_column=i + 1 + lst_merge[i])
        thecell = worksheet[chr(65 + i) + str(line_count + 2)]
        thecell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    lst_title = ['產線別', '區塊數', '總時數', '生產量', '廢料量', '區塊數', '總時數', '生產量', '廢料量', '區塊數', '總時數', '區塊數', '總時數', '廢料量',
                 '區塊數', '總時數', '區塊數', '總時數']
    for i in range(len(lst_title)):
        worksheet.cell(row=line_count + 3, column=i + 1).value = lst_title[i]
        thecell = worksheet[chr(65 + i) + str(line_count + 3)]
        thecell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i in range(sp_data.shape[0]):
        if i < sp_data.shape[0] - 1:
            worksheet.cell(row=line_count + 4 + i, column=1).value = lst_lines[i]
        else:
            worksheet.cell(row=line_count + 4 + i, column=1).value = 'All Lines'
        thecell = worksheet['A' + str(line_count + 4 + i)]
        thecell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for k in range(sp_data.shape[1]):
            worksheet.cell(row=line_count + 4 + i, column=k + 2).value = sp_data[i, k]
            if k == 3 or k == 7:  worksheet.cell(row=line_count + 4 + i, column=k + 2).number_format = u'0.00'
            thecell = worksheet[chr(65 + k + 1) + str(line_count + 4 + i)]
            thecell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 產品需求平衡表
    df_DemandBalanceSheet, df_negStock = DemandBalanceSheet(demand_start, demand_end, lst_weekday, lst_block, df_orderdata, df_products, df_lines)
    df_DemandBalanceSheet.to_excel(writer, sheet_name='產品需求平衡表', startrow=0, startcol=0, index=False)
    worksheet = writer.sheets['產品需求平衡表']
    for i in range(df_DemandBalanceSheet.shape[0] + 1):
        for k in range(df_DemandBalanceSheet.shape[1]):
            worksheet.cell(row=i + 1, column=k + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 逾期生產的產品

    # df_negStock.to_excel(writer, sheet_name='逾期生產的產品', startrow=0, startcol=0, index=False)
    if df_negStock.shape[0] > 0: DrawOnStock(df_negStock, writer, thin)

    # 實施區段圖
    # n_day = 0
    df_sector, df_LRation, lst_PHData = ImpSector_Sheet(lst_block, lst_weekday, max_linearray, begin_day, end_day, maxline_count, df_lines)

    workbook = writer.book
    worksheet = workbook.create_sheet('實施區段圖', 2 if df_order_lost.shape[0] == 0 else 3)
    writer.sheets['實施區段圖'] = worksheet

    # row_count = 0
    with open('imp_sector.txt') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        row_count = sum(1 for row in csv_reader)
    row_count -= 1

    thin = Side(border_style='thin', color='000000')
    line_count = 0
    with open('imp_sector.txt') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            line_count += 1
            for i in range(len(row)):
                celldata = row[i]
                if celldata.find('%%') > 0: worksheet.cell(row=line_count, column=i + 1).alignment = Alignment(wrapText=True)
                worksheet.cell(row=line_count, column=i + 1).value = celldata.replace('%%', '\n')

                if line_count == 2 and celldata != '': worksheet.cell(row=line_count, column=i + 1).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                if line_count < 3 or line_count == row_count or i == 0:
                    worksheet.cell(row=line_count, column=i + 1).alignment = Alignment(horizontal='center', vertical='center')
                else:
                    worksheet.cell(row=line_count, column=i + 1).alignment = Alignment(horizontal='left', vertical='top')
                worksheet.cell(row=line_count, column=i + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    worksheet.cell(row=1, column=1).value = '線別'
    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    worksheet.cell(row=1, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    worksheet.cell(row=2, column=1).fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')

    df_sector.to_excel(writer, sheet_name='實施區段圖', startrow=line_count + 1, startcol=0, index=False)
    df_LRation.to_excel(writer, sheet_name='實施區段圖', startrow=df_sector.shape[0] + line_count + 3, startcol=0, index=False)

    # Draw PH hours on two y-axes.
    df_PHData = pd.DataFrame(columns=[])
    xi = 0
    for x_PH in lst_PHData:
        lst_tmp = x_PH.split(',')
        lst_tmp = [0 if x == '' else (x if not x.replace('.', '').isnumeric() else float(x)) for x in lst_tmp]
        df_PHData[lst_tmp[0]] = lst_tmp[1:]
        xi += 1

    df_PHData.to_excel(writer, sheet_name='實際生產時數', startrow=0, startcol=0, index=False)
    worksheet = writer.sheets['實際生產時數']
    dates = Reference(worksheet, min_row=2, max_row=df_PHData.shape[0] + 1, min_col=1)

    c1 = BarChart()
    v1 = Reference(worksheet, min_col=3, min_row=1, max_row=df_PHData.shape[0] + 1)
    c1.add_data(v1, titles_from_data=True)
    c1.set_categories(dates)
    c1.x_axis.title = df_PHData.columns[0]
    c1.y_axis.title = df_PHData.columns[2] + ' (hrs)'
    c1.y_axis.majorGridlines = None
    c1.y_axis.scaling.max = 30
    c1.y_axis.majorUnit = 6
    c1.title = '實際生產時數趨勢圖'

    # Create a second chart
    c2 = LineChart()
    v2 = Reference(worksheet, min_col=2, min_row=1, max_row=df_PHData.shape[0] + 1)
    c2.add_data(v2, titles_from_data=True)
    c2.set_categories(dates)
    c2.y_axis.axId = 200
    c2.y_axis.scaling.max = 120
    c2.y_axis.majorUnit = 24
    c2.y_axis.title = df_PHData.columns[1] + ' (hrs)'

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c1.y_axis.crosses = 'max'
    c1 += c2

    c1.width = 21
    c1.height = 14
    worksheet.add_chart(c1, chr(66 + df_PHData.shape[1]) + '2')

    lstsheetorder = workbook.sheetnames
    sh_id1 = Opt_func.list_index(lstsheetorder, '實施區段圖')
    sh_id2 = Opt_func.list_index(lstsheetorder, '實際生產時數')
    lstsheetorder = [i for i in range(len(lstsheetorder))]
    lstsheetorder.remove(sh_id2)
    lstsheetorder.insert(sh_id1 + 1, sh_id2)
    workbook._sheets = [workbook._sheets[i] for i in lstsheetorder]

    writer.save()